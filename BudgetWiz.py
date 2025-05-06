import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from BudgetUtils import (  # Import specific functions for better readability
    load_categories,
    clean_text,
    name_transaction,
    categorize_transaction,
    split_transaction,
    file_input,
    print_confirmation,
    open_excel_file,
    DATA_FOLDER, 
    EXCEL_FILE
)


def process_transactions(input_file, sheet_name):
    """Process transactions from a CSV and output to an Excel file with categories and a pivot table."""

    ############################################
    ### DF CLEANING / TRANSACTION PROCESSING ###
    ############################################

    df = pd.read_csv(input_file)

    # Credit/Debit
    df["Credit"] = df["Credit"].astype(float) * -1 # Make Credit negative
    df['Amount'] = df['Debit'].fillna(df['Credit'])

    df = df[~df["Category"].str.lower().str.contains("payment")]  # Remove Payment records

    # Drop Cols
    df = df.drop(columns=["Posted Date", "Card No.", "Debit", "Credit", "Category"])

    # Rename Cols
    df.rename(columns={
        'Description': 'Name',
        'Transaction Date': 'Date',
    }, inplace=True)
    
    # Drop the "Memo" column if it exists
    if "Memo" in df.columns:
        df.drop(columns=["Memo"], inplace=True)

    # Split Name col
    #df[["Name", "City", "State"]] = df["Name"].apply(split_transaction)

    # Clean Names
    df["Name"] = df["Name"].apply(clean_text)


    ### Regex to determine instructions
    if categories:  # Check if categories is NOT empty
        pattern = "|".join(re.escape(k) for k in categories.keys())  # Escape special regex chars
        pattern = "(?i)" + pattern  # Make case-insensitive

        # Identify names that are NOT in categories (case-insensitive)
        new_names = df["Name"][~df["Name"].str.contains(pattern)].unique()
    else:
        new_names = df["Name"].unique()  # All names are "new" if categories is empty


    ### Name transactions
    if new_names.size > 0:  # Print instructions if new names were found
        print(f"\n\U00002728 Edit Transaction Names in {sheet_name}. Press Enter to keep original name.")
    else:
        print(f"\n\U0001F44D All Transactions Recognized in {sheet_name}")
    
    # Apply name_transaction iteratively over input transactions
    df["Name"] = df["Name"].apply(name_transaction)

    ### Apply categorizations
    if new_names.size > 0:  # Print instructions based on new_names value
        print(f"\n\U00002728 Edit Transaction Categories in {sheet_name}. Press Enter to leave uncategorized.")
    
    # Apply categorize_transaction iteratively over input transactions
    df["Category"] = df["Name"].apply(categorize_transaction) 

    # Replace the "Name" column in the transaction table with recognized names from the newly updated categories
    df["Name"] = df["Name"].apply(lambda name: next((key for key in categories if key.lower() in name.lower()), name))

    # Reorder transaction table cols
    df = df[["Date", "Name", "Amount", "Category"]]


    ### Pivot Tables
    # Create a pivot table summing expenses by category
    pivot_total = df.groupby("Category")["Amount"].sum().reset_index()
    pivot_total.sort_values("Amount", ascending= False, inplace=True)

    # Create a pivot table displaying expense frequencies
    pivot_freq = df.groupby("Category")["Amount"].count().reset_index()
    pivot_freq.sort_values("Amount", ascending= False, inplace=True)



    ######################
    ### WRITE TO EXCEL ###
    ######################

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()
    
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(title=sheet_name)
    
    # Write headers without formatting
    headers = list(df.columns)
    ws.append([])  # Blank row for spacing
    ws.append([""] + headers)

    # Determine pivot table start row with a blank space
    pivot_start_row = 26
    pivot_start_col = len(df.columns) + 3
    
    # Write pivot_total header without formatting
    ws.cell(row=pivot_start_row, column=pivot_start_col, value="Category")
    ws.cell(row=pivot_start_row, column=pivot_start_col + 1, value="Amount")

    # Write pivot_freq header without formatting
    ws.cell(row=pivot_start_row, column=pivot_start_col + 3, value="Category")
    ws.cell(row=pivot_start_row, column=pivot_start_col + 4, value="Count")
    
    # Write transactions to Excel
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        for c_idx, value in enumerate(row, start=2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Write pivot_total table
    for r_idx, row in enumerate(pivot_total.itertuples(index=False), start=pivot_start_row + 1):
        for c_idx, value in enumerate(row, start=pivot_start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Write pivot_freq table
    for r_idx, row in enumerate(pivot_freq.itertuples(index=False), start=pivot_start_row + 1):
        for c_idx, value in enumerate(row, start=pivot_start_col + 3):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Write the Grand Total row after the pivot_total table
    ws.cell(row= pivot_start_row + len(pivot_total) + 2, column=pivot_start_col, value="Grand Total")
    ws.cell(row= pivot_start_row + len(pivot_total) + 2, column=pivot_start_col + 1, value= pivot_total["Amount"].sum())

    # Write the Total Count row after the pivot_freq table
    ws.cell(row= pivot_start_row + len(pivot_total) + 2, column=pivot_start_col + 3, value="Transaction Count")
    ws.cell(row= pivot_start_row + len(pivot_total) + 2, column=pivot_start_col + 4, value= pivot_freq["Amount"].sum())



    ##################
    ### FORMATTING ###
    ##################

    ### Border Styles
    all_borders = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    side_borders = Border(left=Side(style='thin'), right=Side(style='thin'))
    bottom_borders = Border(bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
    
    ### Transaction Table Borders
    # Apply all borders to the header row (transaction table)
    for col_idx in range(2, len(headers) + 2):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = all_borders

    # Apply side borders for each column in the transaction table
    last_row = len(df) + 2  # +1 for header row
    for col_idx in range(2, len(headers) + 2):
        for r_idx in range(3, last_row):  # from first data row to last
            ws.cell(row=r_idx, column=col_idx).border = side_borders

    # Apply bottom border to the last row of the transaction table
    for col_idx in range(2, len(headers) + 2):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.border = bottom_borders

    ### Sum Pivot Table Borders
    # Apply all borders to the pivot table header row
    for col_idx in range(pivot_start_col, pivot_start_col + 2):  # Two columns in the pivot table header
        cell = ws.cell(row=pivot_start_row, column=col_idx)
        cell.border = all_borders

    # Apply side borders for each column in the pivot table
    for r_idx in range(pivot_start_row + 1, pivot_start_row + len(pivot_total) + 1):
        for col_idx in range(pivot_start_col, pivot_start_col + 2):  # Two columns in the pivot table
            ws.cell(row=r_idx, column=col_idx).border = side_borders

    # Apply bottom border to the last row of the pivot table (without adding a blank row)
    last_pivot_row = pivot_start_row + len(pivot_total)
    for col_idx in range(pivot_start_col, pivot_start_col + 2):  # Two columns in the pivot table
        cell = ws.cell(row=last_pivot_row, column=col_idx)
        cell.border = bottom_borders

    ### Freq Pivot Table Borders
    # Apply all borders to the pivot table header row
    for col_idx in range(pivot_start_col + 3, pivot_start_col + 5):  # Two columns in the pivot table header
        cell = ws.cell(row=pivot_start_row, column=col_idx)
        cell.border = all_borders

    # Apply side borders for each column in the pivot table
    for r_idx in range(pivot_start_row + 1, pivot_start_row + len(pivot_total) + 1):
        for col_idx in range(pivot_start_col + 3, pivot_start_col + 5):  # Two columns in the pivot table
            ws.cell(row=r_idx, column=col_idx).border = side_borders

    # Apply bottom border to the last row of the pivot table (without adding a blank row)
    last_pivot_row = pivot_start_row + len(pivot_total)
    for col_idx in range(pivot_start_col + 3, pivot_start_col + 5):  # Two columns in the pivot table
        cell = ws.cell(row=last_pivot_row, column=col_idx)
        cell.border = bottom_borders

    ### Grand Total Borders
    # Apply borders to the Grand Total cells
    grand_total_row_index = pivot_start_row + len(pivot_total) + 2  # Row index for the Grand Total row
    for col_idx in range(pivot_start_col, pivot_start_col + 2): 
        cell = ws.cell(row=grand_total_row_index, column=col_idx)
        cell.border = all_borders  # Apply the defined border style
    
    ### Transaction Count Borders
    # Apply borders to the Transaction Count cells
    grand_total_row_index = pivot_start_row + len(pivot_freq) + 2  # Row index for the Grand Total / Transaction Ct. row
    for col_idx in range(pivot_start_col + 3, pivot_start_col + 5): 
        cell = ws.cell(row=grand_total_row_index, column=col_idx)
        cell.border = all_borders  # Apply the defined border style


    ### Bold & Center
    # Define font and alignment styles
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    # Apply bold & center formatting to transaction headers
    for col_idx in range(2, len(headers) + 2):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align

    # Apply bold & center formatting to total_pivot table headers
    for col_idx in range(pivot_start_col, pivot_start_col + 3):  # Two columns in the pivot table
        cell = ws.cell(row=pivot_start_row, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
    
    # Apply bold & center formatting to freq_pivot table headers
    for col_idx in range(pivot_start_col + 3, pivot_start_col + 5):  # Two columns in the pivot table
        cell = ws.cell(row=pivot_start_row, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
    
    # Bold and center the "Grand Total"
    grand_total_cell = ws.cell(row=pivot_start_row + len(pivot_total) + 2, column=pivot_start_col)
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.alignment = Alignment(horizontal="center")

    # Bold and center the "Transaction Count"
    grand_total_cell = ws.cell(row=pivot_start_row + len(pivot_total) + 2, column=pivot_start_col + 3)
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.alignment = Alignment(horizontal="center")


    ### Accounting Format
    # Define accounting format for excel
    accounting_format = '"$"* #,##0.00_);[Red]"$"* #,##0.00;"-";@'
    # Apply format to the amounts in the DataFrame
    for r_idx in range(2, len(df) + 3):  # For transaction table col E
        ws.cell(row=r_idx, column=5).number_format = accounting_format  # Col 4 = E for Amount

    # Apply format to the total_pivot table amounts
    for r_idx in range(pivot_start_row + 1, pivot_start_row + len(pivot_total) + 3):  # For total_pivot table Amount col
        ws.cell(row=r_idx, column=11).number_format = accounting_format  # Column 11 = K for Amount

    
    ### Autofit/Adjust Col/Row Size
    for col in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        adjusted_width = (max_length + 2) * 1.08 #padding
        if adjusted_width > 0:
            ws.column_dimensions[col_letter].width = adjusted_width


    # Adjust column widths and row height
    Adjustment = 0.71 # Openpyxl is not exact
    ws.column_dimensions['E'].width = 8 + Adjustment # Amount col
    ws.column_dimensions['K'].width = 10.14 + Adjustment # Amount col (pivot)

    ws.column_dimensions['I'].width = 2 + Adjustment # between transaction and pivot tables
    ws.column_dimensions['L'].width = 3 + Adjustment # between pivot tables

    #ws.column_dimensions['D'].width += 50
    
    ws.column_dimensions['A'].width = 0.83 + Adjustment - 0.09 # first col for aesthetics
    ws.row_dimensions[1].height = 7.5 # first row for aesthetics




    ######################
    ### DOUGHNUT CHART ###
    ######################

    # Create Doughnut Chart
    doughnut_chart = DoughnutChart()

    # Define data range for the chart (from the pivot table)
    data = Reference(ws, min_col=pivot_start_col + 1, min_row=pivot_start_row + 1, max_row=pivot_start_row + len(pivot_total))  # Amount Data
    labels = Reference(ws, min_col=pivot_start_col, min_row=pivot_start_row + 1, max_row=pivot_start_row + len(pivot_total))  # Category Labels

    # Add data and set categories
    doughnut_chart.add_data(data, titles_from_data=False)
    doughnut_chart.set_categories(labels)

    # Chart Customization
    doughnut_chart.title = "Expense Distribution"
    #doughnut_chart.style = 26  # Change this number to the desired style
    doughnut_chart.dLbls = DataLabelList()
    doughnut_chart.dLbls.showPercent = True
    doughnut_chart.dLbls.showCatName = False
    doughnut_chart.dLbls.showVal = True
    doughnut_chart.dLbls.showSerName = False  
    doughnut_chart.dLbls.showLeaderLines = False

    # Slice Coloration
    serie = doughnut_chart.series[0]
    for i, color in enumerate(['92c5f9', '468fea', '6082b6', '8c92ac ', '717EB2', '4963AA']):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        serie.dPt.append(pt)

    # Adjust Doughnut Hole Size
    doughnut_chart.firstSliceAng = 45  # Smallest slice location by angle (bigger counter-clockwise)
    doughnut_chart.holeSize = 50

    # Set the chart size (increase width and height)
    doughnut_chart.height, doughnut_chart.width = 12, 15

    # Add the chart to the worksheet
    chart_cell = "G2"
    ws.add_chart(doughnut_chart, chart_cell)



    ######################
    ### SAVE AND CLOSE ###
    ######################

    try:
        wb.active = wb.sheetnames.index(sheet_name)
        wb.save(EXCEL_FILE)
        wb.close()
        print(f"\n\U00002705 Processed transactions successfully saved to sheet '{sheet_name}' in: \n{EXCEL_FILE}\n")
    except Exception as e:  # Catch a broad exception, or be more specific
        print(f"\n\U0000274C Error saving transactions: {e}\n")  # Red X for error
        # Optionally, you might want to rollback any changes, log the error, etc.

    



if __name__ == "__main__":
    # Loading Files/Paths
    categories = load_categories()
    script_dir = os.path.dirname(os.path.abspath(__file__)) # BudgetWiz.py directory path
    
    # Brief Message
    print("-" * 70)
    print(f"\U0001F4A1 Note: All user inputs are case-insensitive unless otherwise noted")
    
    # File I/O
    files_and_sheets = file_input() # File Input
    print_confirmation(files_and_sheets) # Print File Confirmation
    
    # Process each CSV and create corresponding sheets
    for csv_file, sheet_name in files_and_sheets:
        process_transactions(csv_file, sheet_name)
    
    open_excel_file() # Open XLSX File